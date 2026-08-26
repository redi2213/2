# -*- coding: utf-8 -*-
"""
ابزار تولید گزارش متنی خاموشی‌های خطوط ۲۰ کیلوولت
---------------------------------------------------
این ابزار فایل اکسل قطع و وصل خطوط را می‌خواند، بازه‌ای از روزها را از کاربر
می‌پرسد و برای هر خاموشی در آن بازه یک پاراگراف متنی می‌سازد؛ سپس همه را در
یک فایل txt ذخیره می‌کند.

فقط این ۶ شیت پردازش می‌شوند (شیت‌های خطوط ۲۰ کیلوولت):
  کیشکور، علی آباد   -> دیسپاچینگ توزیع ایرانشهر
  سروش، مورتان، ستایش، هاتف -> دیسپاچینگ توزیع چابهار

ساختار ستون‌ها در هر شیت (پس از بازچینی) از ردیف ۹ به بعد:
  A=ردیف  B=روز  C=ساعت قطع  D=ساعت وصل  E=مدت قطع
  F=آمپراژ قطع  G=آمپراژ وصل  H=علت قطع  I=قطع حفاظتی  J=قطع دستی  K=نمراتور
نام خط: C5   |   سال: C6   |   ماه: G6
"""

import sys
import os
import openpyxl

# ---------------------------------------------------------------------------
# پیکربندی: نام شیت‌ها و دیسپاچینگ مربوطه
# ---------------------------------------------------------------------------
LINE_SHEETS = [
    ("قطع و وصل خط  کیشکور", "ایرانشهر"),
    ("قطع و وصل خط علی آباد", "ایرانشهر"),
    ("قطع و وصل خط سروش", "چابهار"),
    ("قطع و وصل خط مورتان", "چابهار"),
    ("قطع و وصل خط ستایش", "چابهار"),
    ("قطع و وصل خط هاتف ", "چابهار"),
]

HEADER_ROW = 8
DATA_START_ROW = 9

# نگاشت علت قطع دراپ‌داون -> عبارتی که در متن نوشته می‌شود (بعد از "به علت ")
CAUSE_TEXT_MAP = {
    "مدیریت بار": "کمبود تولید",
    # بقیه‌ی گزینه‌ها عیناً همان متن دراپ‌داون نوشته می‌شوند و نیازی به نگاشت ندارند:
    # اورکارنت لحظه‌ای, اورکارنت زمانی, ارت فالت لحظه‌ای, ارت فالت زمانی,
    # ارت فالت نقطه حساس, مانور شبکه, تعمیرات شبکه, ترمیم رابط, شکستن تیر برق,
    # اورکارنت و ارت فالت همزمان, صاعقه, اضافه بار, دوفاز شدن
}

PERSIAN_MONTHS = [
    "فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور",
    "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند",
]

PERSIAN_DIGITS = "۰۱۲۳۴۵۶۷۸۹"
LATIN_DIGITS = "0123456789"
_DIGIT_MAP = str.maketrans(PERSIAN_DIGITS, LATIN_DIGITS)


def to_latin_digits(value):
    """تبدیل ارقام فارسی به لاتین در صورت وجود، برای مقایسه و فرمت امن."""
    if value is None:
        return value
    return str(value).translate(_DIGIT_MAP)


def normalize_time(raw):
    """
    ساعت‌ها در اکسل به صورت عدد صحیح مثل 1145 یا 900 ذخیره شده‌اند (ساعت:دقیقه چسبیده).
    این تابع آن‌ها را به رشته‌ی "HH:MM" تبدیل می‌کند.
    """
    if raw is None or raw == "":
        return None
    s = to_latin_digits(raw).strip()
    if s == "":
        return None
    try:
        num = int(float(s))
    except ValueError:
        return s  # اگر از قبل رشته‌ی متنی دلخواه بود، همان را برگردان
    hour = num // 100
    minute = num % 100
    return f"{hour:02d}:{minute:02d}"


def normalize_amp(raw):
    """آمپراژ را به رشته‌ی تمیز (بدون اعشار غیرضروری) تبدیل می‌کند، یا None اگر خالی باشد."""
    if raw is None or raw == "":
        return None
    s = to_latin_digits(raw).strip()
    if s == "":
        return None
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
        return str(f)
    except ValueError:
        return s


def time_sort_key(time_str):
    """کلید مرتب‌سازی بر اساس ساعت قطع (HH:MM -> عدد دقیقه از نیمه‌شب)."""
    if not time_str:
        return 10 ** 9  # ردیف‌های بدون ساعت را انتهای لیست بگذار
    try:
        h, m = time_str.split(":")
        return int(h) * 60 + int(m)
    except Exception:
        return 10 ** 9


def cause_phrase(raw_cause):
    """عبارت نهایی علت قطع برای درج بعد از 'به علت '."""
    if raw_cause is None:
        return None
    cause = str(raw_cause).strip()
    if cause == "":
        return None
    return CAUSE_TEXT_MAP.get(cause, cause)


def build_sentence(line_name, dispatch, cause, off_time, off_amp, on_time, on_amp):
    """دو خط متن مربوط به یک خاموشی را می‌سازد (خروج از مدار + ورود به مدار)."""
    lines = []
    off_sentence = (
        f"خط {line_name} به علت {cause} به درخواست دیسپاچینگ توزیع {dispatch} "
        f"در ساعت {off_time} با {off_amp} آمپر از مدار خارج گردید."
    )
    lines.append(off_sentence)

    if on_time:
        if on_amp:
            on_sentence = (
                f"خط {line_name} به درخواست دیسپاچینگ توزیع {dispatch} "
                f"در ساعت {on_time} با {on_amp} آمپر وارد مدار گردید."
            )
        else:
            on_sentence = (
                f"خط {line_name} به درخواست دیسپاچینگ توزیع {dispatch} "
                f"در ساعت {on_time} وارد مدار گردید."
            )
        lines.append(on_sentence)

    return lines


def collect_outages(wb, day_from, day_to):
    """
    تمام خاموشی‌های شیت‌های خط را در بازه‌ی روز مشخص جمع‌آوری می‌کند.
    خروجی: دیکشنری {day: [outage_dict, ...]}
    هر outage_dict شامل: line_name, dispatch, cause, off_time, off_amp, on_time, on_amp
    """
    outages_by_day = {}

    for sheet_name, dispatch in LINE_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"⚠ هشدار: شیت «{sheet_name}» در فایل پیدا نشد، رد می‌شود.")
            continue

        ws = wb[sheet_name]
        line_name = ws["C5"].value
        line_name = str(line_name).strip() if line_name else sheet_name

        # پیدا کردن آخرین ردیف داده (تا قبل از ردیف "تعداد خاموشی‌ها")
        r = DATA_START_ROW
        last_row = DATA_START_ROW - 1
        while True:
            day_val = ws[f"B{r}"].value
            # وقتی به ردیف خلاصه (تعداد خاموشی‌ها) یا انتهای جدول برسیم متوقف شو.
            label_cell = ws[f"A{r}"].value
            if isinstance(label_cell, str) and "تعداد خاموشی" in label_cell:
                break
            if day_val is None and r > DATA_START_ROW + 60:
                break
            if r > DATA_START_ROW + 60:
                break
            last_row = r
            r += 1

        for row in range(DATA_START_ROW, last_row + 1):
            day_raw = ws[f"B{row}"].value
            if day_raw is None or str(day_raw).strip() == "":
                continue
            try:
                day_num = int(to_latin_digits(day_raw))
            except (ValueError, TypeError):
                continue

            if day_num < day_from or day_num > day_to:
                continue

            off_time = normalize_time(ws[f"C{row}"].value)
            on_time = normalize_time(ws[f"D{row}"].value)
            off_amp = normalize_amp(ws[f"F{row}"].value)
            on_amp = normalize_amp(ws[f"G{row}"].value)
            cause = cause_phrase(ws[f"H{row}"].value)

            if off_time is None or off_amp is None or cause is None:
                # داده‌ی ناقص برای خروج از مدار؛ از این ردیف صرف‌نظر کن.
                continue

            outage = {
                "line_name": line_name,
                "dispatch": dispatch,
                "cause": cause,
                "off_time": off_time,
                "off_amp": off_amp,
                "on_time": on_time,
                "on_amp": on_amp,
            }
            outages_by_day.setdefault(day_num, []).append(outage)

    return outages_by_day


def get_month_year(wb):
    """سال و ماه را از اولین شیت خط موجود می‌خواند (C6 و G6)."""
    for sheet_name, _ in LINE_SHEETS:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            year = ws["C6"].value
            month = ws["G6"].value
            if year and month:
                return to_latin_digits(year), str(month).strip()
    return None, None


def format_report(outages_by_day, year, month):
    """متن نهایی گزارش را می‌سازد."""
    out_lines = []
    for day_num in sorted(outages_by_day.keys()):
        day_outages = outages_by_day[day_num]
        day_outages.sort(key=lambda o: time_sort_key(o["off_time"]))

        day_str = f"{int(day_num):02d}"
        out_lines.append(f"تاریخ {year}/{month_number(month):02d}/{day_str}")
        out_lines.append("")

        for idx, o in enumerate(day_outages, start=1):
            sentences = build_sentence(
                o["line_name"], o["dispatch"], o["cause"],
                o["off_time"], o["off_amp"], o["on_time"], o["on_amp"],
            )
            out_lines.append(f"{idx}. {sentences[0]}")
            for extra in sentences[1:]:
                out_lines.append(extra)
            out_lines.append("")

        out_lines.append("")

    return "\n".join(out_lines).rstrip() + "\n"


def month_number(month_name):
    try:
        return PERSIAN_MONTHS.index(month_name) + 1
    except ValueError:
        return 0


def prompt_int(prompt_text):
    while True:
        raw = input(prompt_text).strip()
        raw = to_latin_digits(raw)
        try:
            return int(raw)
        except ValueError:
            print("لطفاً فقط یک عدد وارد کنید.")


def find_excel_file():
    """فایل اکسل را در همان پوشه‌ی اجرای ابزار پیدا می‌کند."""
    here = os.path.dirname(os.path.abspath(sys.argv[0] if not is_frozen() else sys.executable))
    candidates = [f for f in os.listdir(here) if f.lower().endswith(".xlsx") and not f.startswith("~$")]
    if not candidates:
        return None, here
    if len(candidates) == 1:
        return os.path.join(here, candidates[0]), here
    print("چند فایل اکسل در این پوشه پیدا شد:")
    for i, c in enumerate(candidates, start=1):
        print(f"  {i}) {c}")
    while True:
        choice = input("شماره‌ی فایل موردنظر را وارد کنید: ").strip()
        choice = to_latin_digits(choice)
        try:
            idx = int(choice) - 1
            if 0 <= idx < len(candidates):
                return os.path.join(here, candidates[idx]), here
        except ValueError:
            pass
        print("انتخاب نامعتبر است.")


def is_frozen():
    return getattr(sys, "frozen", False)


def main():
    print("=" * 60)
    print("ابزار تولید گزارش متنی خاموشی خطوط ۲۰ کیلوولت")
    print("=" * 60)

    excel_path, base_dir = find_excel_file()
    if not excel_path:
        print("❌ هیچ فایل اکسلی (.xlsx) کنار این ابزار پیدا نشد.")
        input("برای خروج Enter را بزنید...")
        return

    print(f"فایل اکسل: {os.path.basename(excel_path)}")

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"❌ خطا در باز کردن فایل اکسل: {e}")
        input("برای خروج Enter را بزنید...")
        return

    day_from = prompt_int("روز شروع را وارد کنید (مثلاً 5): ")
    day_to = prompt_int("روز پایان را وارد کنید (مثلاً 10): ")
    if day_to < day_from:
        day_from, day_to = day_to, day_from

    year, month = get_month_year(wb)
    if not year or not month:
        print("❌ نتوانستم سال/ماه را از فایل اکسل بخوانم (سلول‌های C6/G6).")
        input("برای خروج Enter را بزنید...")
        return

    outages_by_day = collect_outages(wb, day_from, day_to)

    if not outages_by_day:
        print(f"در بازه‌ی روز {day_from} تا {day_to} هیچ خاموشی‌ای پیدا نشد.")
        input("برای خروج Enter را بزنید...")
        return

    report_text = format_report(outages_by_day, year, month)

    out_name = f"گزارش خاموشی از {day_from} تا {day_to} - {month} {year}.txt"
    out_path = os.path.join(base_dir, out_name)
    with open(out_path, "w", encoding="utf-8-sig") as f:
        f.write(report_text)

    print()
    print(f"✅ گزارش ساخته شد: {out_name}")
    print(f"مسیر: {out_path}")
    input("برای خروج Enter را بزنید...")


if __name__ == "__main__":
    main()
